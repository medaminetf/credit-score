# -*- coding: utf-8 -*-
"""
Notation AFRICAPITAL des émetteurs de dette privée
---------------------------------------------------

Fonctionnement :
1. Le fichier "Liste_Emetteurs (2).xlsx" est conservé dans le même dossier
   que ce fichier app.py. Il contient le référentiel des émetteurs :
   nom, type et secteur.
2. L'utilisateur charge depuis l'interface Streamlit le fichier de données
   contenant les feuilles "financier" et "corporate".
3. L'application calcule les ratios, les notes et la note globale.

Lancement : streamlit run app.py
"""

import io
import re
import unicodedata
from pathlib import Path

import pandas as pd
import streamlit as st


# ----------------------------------------------------------------------------
# Configuration générale
# ----------------------------------------------------------------------------

st.set_page_config(
    page_title="Notation Émetteurs — Dette Privée",
    page_icon="📊",
    layout="wide",
)

NAVY = "#1F2A44"
GOLD = "#C9A227"

RATING_META = {
    "A": {
        "color": "#16A34A",
        "bg": "#F0FDF4",
        "border": "#86EFAC",
        "label": "Qualité de crédit forte",
    },
    "B": {
        "color": "#2563EB",
        "bg": "#EFF6FF",
        "border": "#93C5FD",
        "label": "Qualité de crédit satisfaisante",
    },
    "C": {
        "color": "#D97706",
        "bg": "#FFFBEB",
        "border": "#FCD34D",
        "label": "Sous surveillance",
    },
    "D": {
        "color": "#DC2626",
        "bg": "#FEF2F2",
        "border": "#FCA5A5",
        "label": "Qualité de crédit dégradée",
    },
    "N/A": {
        "color": "#6B7280",
        "bg": "#F9FAFB",
        "border": "#D1D5DB",
        "label": "Données non disponibles",
    },
}

POINTS = {"A": 4, "B": 3, "C": 2, "D": 1}

st.markdown(
    f"""
    <style>
      .block-container {{ padding-top: 2rem; }}
      .acm-title {{
          text-align: center; font-size: 2.3rem; font-weight: 800;
          color: {NAVY}; margin-bottom: .2rem;
      }}
      .acm-issuer {{ font-size: 1.6rem; font-weight: 700; color: #111827; }}
      .acm-badge {{
          display: inline-block; padding: 2px 10px; border-radius: 8px;
          background: #FEE2E2; color: #B91C1C; font-size: .85rem; font-weight: 600;
          margin-right: 8px;
      }}
      .acm-sector {{ color: #4B5563; font-size: .9rem; }}
      .rating-card {{
          border-radius: 12px; padding: 14px 30px; text-align: center;
          border: 2px solid; min-width: 190px;
      }}
      .rating-letter {{ font-size: 2.2rem; font-weight: 800; line-height: 1.1; }}
      .rating-label {{ font-size: .75rem; color: #374151; }}
      .kpi-label {{ color: #374151; font-size: .9rem; margin-bottom: .1rem; }}
      .kpi-value {{ font-size: 2rem; font-weight: 700; color: #111827; }}
      .kpi-note {{ font-size: .8rem; font-weight: 700; }}
      hr {{ margin: 1.2rem 0; }}
    </style>
    """,
    unsafe_allow_html=True,
)


# ----------------------------------------------------------------------------
# Moteur de notation
# ----------------------------------------------------------------------------


def note_solvabilite(x):
    """Fonds propres / total actif."""
    if x is None:
        return "N/A"
    return "A" if x >= 0.10 else "B" if x >= 0.08 else "C" if x >= 0.06 else "D"


def note_exigible_pnb(x):
    """Total exigible / PNB — un levier faible est meilleur."""
    if x is None:
        return "N/A"
    return "A" if x <= 20 else "B" if x <= 30 else "C" if x <= 40 else "D"


def note_marge_intermediation(x):
    """Produits d'intérêts / charges d'intérêts."""
    if x is None:
        return "N/A"
    return "A" if x >= 2.5 else "B" if x >= 2 else "C" if x >= 1.5 else "D"


def note_gearing(x):
    """Dette nette / fonds propres."""
    if x is None:
        return "N/A"
    return "A" if x <= 0.5 else "B" if x <= 0.7 else "C" if x <= 0.9 else "D"


def note_dn_ebitda(x):
    """Dette nette / EBITDA."""
    if x is None:
        return "N/A"
    return "A" if x <= 1 else "B" if x <= 2 else "C" if x <= 4 else "D"


def note_couverture(x):
    """EBITDA / frais financiers."""
    if x is None:
        return "N/A"
    return "A" if x >= 15 else "B" if x >= 10 else "C" if x >= 5 else "D"


def note_globale(notes):
    """Moyenne des points A=4, B=3, C=2, D=1, reconvertie en lettre."""
    pts = [POINTS[n] for n in notes if n in POINTS]
    if len(pts) < 3:
        return None, "N/A"

    moyenne = sum(pts) / len(pts)
    lettre = (
        "A"
        if moyenne >= 3.5
        else "B"
        if moyenne >= 2.5
        else "C"
        if moyenne >= 1.5
        else "D"
    )
    return round(moyenne, 2), lettre


def _num(v):
    """Convertit une cellule en float. Renvoie None si elle est vide ou non numérique."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None

    try:
        f = float(v)
        return None if f != f else f
    except (TypeError, ValueError):
        return None


def _div(a, b):
    """Division sécurisée."""
    if a is None or b in (None, 0):
        return None
    return a / b


# ----------------------------------------------------------------------------
# Lecture robuste des fichiers Excel
# ----------------------------------------------------------------------------


def _norm(s):
    """Normalise un libellé : sans accents, minuscules, espaces compactés."""
    if s is None:
        return ""

    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    s = s.lower().strip()
    s = re.sub(r"[’']", "'", s)
    s = re.sub(r"\s+", " ", s)
    return s


ALIASES = {
    "societes": ["societes", "societe", "emetteur", "emetteurs"],
    "capitaux_propres": ["capitaux propres", "capitaux propre"],
    "total_actif": ["total actif", "total actifs"],
    "dettes_subordonnees": ["dettes subordonnees", "dette subordonnee"],
    "etablissements_credit": [
        "etablissements de credits",
        "etablissement de credit",
        "etablissements de credit",
    ],
    "dettes_clientele": ["dettes envers la clientele", "dette envers la clientele"],
    "autres_dettes_titre": [
        "autre dettes representees par un titre",
        "autres dettes representees par un titre",
    ],
    "pnb": ["produit net bancaire", "pnb"],
    "produits_interet": ["produits d'interet", "produit d'interet"],
    "charges_interet": [
        "charges d'interets",
        "charge d'interets",
        "charges d'interet",
    ],
    "dette_nette": ["dette nette"],
    "resultat_exploitation": ["resultat d'exploitation", "resultat dexploitation"],
    "ebitda": ["ebitda"],
    "frais_financiers": [
        "charges d'interet (frais financiers)",
        "frais financiers",
        "charges d'interets (frais financiers)",
    ],
    "type": ["type"],
    "secteur": ["secteur"],
}


def _map_columns(columns):
    """Associe chaque clé canonique à son nom de colonne réel."""
    norm_map = {c: _norm(c) for c in columns}
    mapping = {}

    for key, aliases in ALIASES.items():
        aliases_norm = {_norm(a) for a in aliases}
        for original, normalise in norm_map.items():
            if normalise in aliases_norm:
                mapping[key] = original
                break

    return mapping


def _detect_header_row(ws, max_scan=10):
    """Trouve la ligne contenant l'en-tête Société/Émetteur."""
    societes_aliases = {_norm(a) for a in ALIASES["societes"]}

    for numero_ligne in range(1, max_scan + 1):
        for cellule in ws[numero_ligne]:
            if _norm(cellule.value) in societes_aliases:
                return numero_ligne

    return 1


def _find_sheet(sheetnames, keywords):
    """Trouve la première feuille dont le nom contient un mot-clé."""
    for nom in sheetnames:
        nom_normalise = _norm(nom)
        if any(mot in nom_normalise for mot in keywords):
            return nom

    return None


def _lire_referentiel(contenu_base: bytes):
    """Lit le fichier fixe Liste_Emetteurs (2).xlsx présent dans Git."""
    import openpyxl

    wb_base = openpyxl.load_workbook(
        io.BytesIO(contenu_base),
        read_only=True,
        data_only=True,
    )

    sheet_univers = _find_sheet(
        wb_base.sheetnames,
        ["feuil1", "univers", "liste", "emetteur"],
    )

    # Si aucun nom de feuille ne correspond, on utilise la première feuille.
    if sheet_univers is None and wb_base.sheetnames:
        sheet_univers = wb_base.sheetnames[0]

    if sheet_univers is None:
        raise ValueError(
            "Le fichier de base Liste_Emetteurs (2).xlsx ne contient aucune feuille."
        )

    header_u = _detect_header_row(wb_base[sheet_univers])
    xls_base = pd.ExcelFile(io.BytesIO(contenu_base))

    raw_u = pd.read_excel(
        xls_base,
        sheet_name=sheet_univers,
        header=header_u - 1,
    )

    raw_u.columns = [str(c).strip() for c in raw_u.columns]
    raw_u = raw_u.loc[:, ~raw_u.columns.str.startswith("Unnamed")]
    cmap_u = _map_columns(raw_u.columns)

    if "societes" not in cmap_u:
        raise ValueError(
            "La colonne 'Émetteur' ou 'Sociétés' est introuvable dans "
            "Liste_Emetteurs (2).xlsx."
        )

    emetteurs = raw_u[cmap_u["societes"]]

    univers = pd.DataFrame(
        {
            "Emetteur": emetteurs,
            "Type": raw_u[cmap_u["type"]] if "type" in cmap_u else "—",
            "Secteur": raw_u[cmap_u["secteur"]] if "secteur" in cmap_u else "—",
        }
    )

    univers = univers.dropna(subset=["Emetteur"]).copy()
    univers["Emetteur"] = univers["Emetteur"].astype(str).str.strip()
    univers["Type"] = univers["Type"].fillna("—").astype(str).str.strip()
    univers["Secteur"] = univers["Secteur"].fillna("—").astype(str).str.strip()

    univers = univers[
        (univers["Emetteur"].str.len() > 0)
        & (univers["Emetteur"].str.lower() != "nan")
    ]

    univers = univers.drop_duplicates(subset=["Emetteur"], keep="first")

    return univers


@st.cache_data(show_spinner=False)
def charger_donnees(contenu_base: bytes, contenu_data: bytes):
    """
    Lit deux fichiers distincts :
    - contenu_base : référentiel Émetteur / Type / Secteur ;
    - contenu_data : données financières à analyser.
    """
    import openpyxl

    # ------------------------------------------------------------------------
    # 1. Référentiel fixe présent dans Git
    # ------------------------------------------------------------------------
    univers = _lire_referentiel(contenu_base)

    # ------------------------------------------------------------------------
    # 2. Fichier de données chargé par l'utilisateur
    # ------------------------------------------------------------------------
    wb_data = openpyxl.load_workbook(
        io.BytesIO(contenu_data),
        read_only=True,
        data_only=True,
    )

    sheet_fin = _find_sheet(wb_data.sheetnames, ["financ"])
    sheet_corp = _find_sheet(wb_data.sheetnames, ["corporate", "corpo"])

    if sheet_fin is None or sheet_corp is None:
        raise ValueError(
            "Le fichier chargé doit contenir les feuilles 'financier' et "
            f"'corporate'. Feuilles trouvées : {wb_data.sheetnames}"
        )

    xls_data = pd.ExcelFile(io.BytesIO(contenu_data))

    # ------------------------------------------------------------------------
    # Sociétés financières
    # ------------------------------------------------------------------------
    header_f = _detect_header_row(wb_data[sheet_fin])

    fin_raw = pd.read_excel(
        xls_data,
        sheet_name=sheet_fin,
        header=header_f - 1,
    )

    fin_raw.columns = [str(c).strip() for c in fin_raw.columns]
    fin_raw = fin_raw.loc[:, ~fin_raw.columns.str.startswith("Unnamed")]
    cmap_f = _map_columns(fin_raw.columns)

    if "societes" not in cmap_f:
        raise ValueError(
            f"La colonne 'Émetteur' ou 'Sociétés' est introuvable dans la feuille '{sheet_fin}'."
        )

    def g(row, key):
        colonne = cmap_f.get(key)
        return _num(row.get(colonne)) if colonne else None

    fin_rows = []

    for _, ligne in fin_raw.iterrows():
        nom = ligne.get(cmap_f.get("societes"))

        if not isinstance(nom, str) or not nom.strip():
            continue

        cp = g(ligne, "capitaux_propres")
        ta = g(ligne, "total_actif")
        dsub = g(ligne, "dettes_subordonnees")
        ec = g(ligne, "etablissements_credit")
        dcl = g(ligne, "dettes_clientele")
        adt = g(ligne, "autres_dettes_titre")
        pnb = g(ligne, "pnb")
        pi = g(ligne, "produits_interet")
        ci = g(ligne, "charges_interet")

        exigible = None
        parties_exigible = [dsub, ec, dcl, adt]

        if any(partie is not None for partie in parties_exigible):
            exigible = sum(
                partie for partie in parties_exigible if partie is not None
            )

        ratio_1 = _div(cp, ta)
        ratio_2 = _div(exigible, pnb)
        ratio_3 = _div(pi, ci)

        note_1 = note_solvabilite(ratio_1)
        note_2 = note_exigible_pnb(ratio_2)
        note_3 = note_marge_intermediation(ratio_3)

        moyenne, lettre = note_globale([note_1, note_2, note_3])

        fin_rows.append(
            {
                "Emetteur": nom.strip(),
                "Famille": "Société financière",
                "Ratio 1": ratio_1,
                "Ratio 2": ratio_2,
                "Ratio 3": ratio_3,
                "Note 1": note_1,
                "Note 2": note_2,
                "Note 3": note_3,
                "Score": moyenne,
                "Note finale": lettre,
            }
        )

    df_fin = pd.DataFrame(fin_rows)

    # ------------------------------------------------------------------------
    # Corporates
    # ------------------------------------------------------------------------
    header_c = _detect_header_row(wb_data[sheet_corp])

    corp_raw = pd.read_excel(
        xls_data,
        sheet_name=sheet_corp,
        header=header_c - 1,
    )

    corp_raw.columns = [str(c).strip() for c in corp_raw.columns]
    corp_raw = corp_raw.loc[:, ~corp_raw.columns.str.startswith("Unnamed")]
    cmap_c = _map_columns(corp_raw.columns)

    if "societes" not in cmap_c:
        raise ValueError(
            f"La colonne 'Émetteur' ou 'Sociétés' est introuvable dans la feuille '{sheet_corp}'."
        )

    def gc(row, key):
        colonne = cmap_c.get(key)
        return _num(row.get(colonne)) if colonne else None

    corp_rows = []

    for _, ligne in corp_raw.iterrows():
        nom = ligne.get(cmap_c.get("societes"))

        if not isinstance(nom, str) or not nom.strip():
            continue

        cp = gc(ligne, "capitaux_propres")
        dette_nette = gc(ligne, "dette_nette")
        resultat_exploitation = gc(ligne, "resultat_exploitation")
        ebitda = gc(ligne, "ebitda")
        frais_financiers = gc(ligne, "frais_financiers")

        # Si l'EBITDA est absent, on utilise le résultat d'exploitation.
        if ebitda is None:
            ebitda = resultat_exploitation

        ratio_1 = _div(dette_nette, cp)
        ratio_2 = _div(dette_nette, ebitda)
        ratio_3 = _div(ebitda, frais_financiers)

        note_1 = note_gearing(ratio_1)
        note_2 = note_dn_ebitda(ratio_2)
        note_3 = note_couverture(ratio_3)

        moyenne, lettre = note_globale([note_1, note_2, note_3])

        corp_rows.append(
            {
                "Emetteur": nom.strip(),
                "Famille": "Corporate",
                "Ratio 1": ratio_1,
                "Ratio 2": ratio_2,
                "Ratio 3": ratio_3,
                "Note 1": note_1,
                "Note 2": note_2,
                "Note 3": note_3,
                "Score": moyenne,
                "Note finale": lettre,
            }
        )

    df_corp = pd.DataFrame(corp_rows)

    return univers, df_fin, df_corp


# ----------------------------------------------------------------------------
# Libellés et formatage
# ----------------------------------------------------------------------------

LIBELLES = {
    "Société financière": [
        ("Ratio de solvabilité", "pct"),
        ("Total exigible / PNB", "x"),
        ("Produits d'intérêts / Charges d'intérêts", "x"),
    ],
    "Corporate": [
        ("Gearing (Dette nette / Fonds propres)", "x"),
        ("Dette nette / EBITDA", "x"),
        ("EBITDA / Frais financiers", "x"),
    ],
}


def fmt(v, kind):
    if v is None or pd.isna(v):
        return "—"

    if kind == "pct":
        return f"{v * 100:.1f}%".replace(".", ",")

    return f"{v:.1f}x".replace(".", ",")


# ----------------------------------------------------------------------------
# Interface Streamlit
# ----------------------------------------------------------------------------

# Nom exact du fichier de base conservé dans Git.
FICHIER_BASE = Path(__file__).parent / "Liste_Emetteurs (2).xlsx"

with st.sidebar:
    st.header("Sélection")

    fichier_data = st.file_uploader(
        "Charger le fichier Data-Emetteurs",
        type=["xlsx"],
        help=(
            "Ce fichier doit contenir les feuilles 'financier' et 'corporate'. "
            "Le référentiel Émetteur / Type / Secteur est lu automatiquement "
            "depuis Liste_Emetteurs (2).xlsx."
        ),
    )

# Vérification du fichier fixe présent dans le dépôt Git.
if not FICHIER_BASE.exists():
    st.error(
        "Le fichier de base 'Liste_Emetteurs (2).xlsx' est introuvable. "
        "Placez-le dans le même dossier que app.py."
    )
    st.stop()

contenu_base = FICHIER_BASE.read_bytes()

# Le fichier de données doit être chargé par l'utilisateur.
if fichier_data is None:
    st.info(
        "Le référentiel des émetteurs est prêt. Chargez maintenant le fichier "
        "Data-Emetteurs.xlsx dans la barre latérale pour lancer l'analyse."
    )
    st.stop()

contenu_data = fichier_data.getvalue()

try:
    univers, df_fin, df_corp = charger_donnees(
        contenu_base,
        contenu_data,
    )
except Exception as erreur:
    st.error(f"Impossible de lire les fichiers : {erreur}")
    st.stop()

# Assemblage des résultats calculés.
frames_non_vides = [df for df in [df_fin, df_corp] if not df.empty]

if not frames_non_vides:
    st.warning("Aucune donnée exploitable n'a été trouvée dans le fichier chargé.")
    st.stop()

resultats = pd.concat(frames_non_vides, ignore_index=True)


# ----------------------------------------------------------------------------
# Jointure entre résultats et référentiel Type / Secteur
# ----------------------------------------------------------------------------

univers = univers.copy()
univers["Cle_emetteur"] = univers["Emetteur"].map(_norm)

# Évite les doublons d'index dans le référentiel.
univers_idx = (
    univers.drop_duplicates(subset=["Cle_emetteur"], keep="first")
    .set_index("Cle_emetteur")
)


def info_univers(nom):
    cle = _norm(nom)

    # Correspondance exacte.
    if cle in univers_idx.index:
        ligne_univers = univers_idx.loc[cle]
        return ligne_univers.get("Type", "—"), ligne_univers.get("Secteur", "—")

    # Correspondance partielle dans les deux sens.
    correspondances = univers_idx[
        univers_idx.index.to_series().apply(
            lambda valeur: cle in valeur or valeur in cle
        )
    ]

    if correspondances.empty:
        return "—", "—"

    ligne_univers = correspondances.iloc[0]
    return ligne_univers.get("Type", "—"), ligne_univers.get("Secteur", "—")


resultats[["Type", "Secteur"]] = resultats["Emetteur"].apply(
    lambda nom: pd.Series(info_univers(nom))
)


# ----------------------------------------------------------------------------
# Filtres de la barre latérale
# ----------------------------------------------------------------------------

with st.sidebar:
    types_dispo = sorted(
        type_emetteur
        for type_emetteur in univers["Type"].dropna().unique().tolist()
        if str(type_emetteur).strip()
    )

    types_sel = st.multiselect(
        "Type d'émetteur",
        types_dispo,
        default=types_dispo,
    )

    if types_sel:
        disponibles = resultats[resultats["Type"].isin(types_sel)]
    else:
        disponibles = resultats

    notes_ok = disponibles[disponibles["Note finale"] != "N/A"]
    choix = notes_ok["Emetteur"].tolist() or disponibles["Emetteur"].tolist()

    emetteur = st.selectbox("Émetteur", choix) if choix else None


# ----------------------------------------------------------------------------
# Affichage principal
# ----------------------------------------------------------------------------

st.markdown(
    '<div class="acm-title">Notation AFRICAPITAL des émetteurs de dette privée</div>',
    unsafe_allow_html=True,
)
st.markdown("---")

if emetteur is None:
    st.warning("Aucun émetteur disponible pour cette sélection.")
    st.stop()

ligne = resultats[resultats["Emetteur"] == emetteur].iloc[0]
note = ligne["Note finale"] if pd.notna(ligne["Note finale"]) else "N/A"
meta = RATING_META.get(note, RATING_META["N/A"])

# En-tête émetteur et carte de notation.
c1, c2 = st.columns([3, 1])

with c1:
    st.markdown(
        f'<div class="acm-issuer">{emetteur}</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        f'<span class="acm-badge">{ligne["Type"]}</span>'
        f'<span class="acm-sector">Secteur : {ligne["Secteur"]}</span>',
        unsafe_allow_html=True,
    )

with c2:
    st.markdown(
        f"""
        <div class="rating-card"
             style="background:{meta['bg']}; border-color:{meta['border']};">
          <div class="rating-letter" style="color:{meta['color']};">{note}</div>
          <div class="rating-label">{meta['label']}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

if ligne["Score"] is not None and pd.notna(ligne["Score"]):
    st.caption(f"Score moyen : {str(ligne['Score']).replace('.', ',')} / 4")


# ----------------------------------------------------------------------------
# Indicateurs clés
# ----------------------------------------------------------------------------

st.markdown("### Indicateurs clés")
libelles = LIBELLES[ligne["Famille"]]

colonnes = st.columns(3)

for colonne, (libelle, format_ratio), cle_ratio, cle_note in zip(
    colonnes,
    libelles,
    ["Ratio 1", "Ratio 2", "Ratio 3"],
    ["Note 1", "Note 2", "Note 3"],
):
    note_ratio = ligne[cle_note]
    couleur = RATING_META.get(note_ratio, RATING_META["N/A"])["color"]

    with colonne:
        st.markdown(
            f"""
            <div class="kpi-label">{libelle}</div>
            <div class="kpi-value">{fmt(ligne[cle_ratio], format_ratio)}</div>
            <div class="kpi-note" style="color:{couleur};">Note : {note_ratio}</div>
            """,
            unsafe_allow_html=True,
        )

st.markdown("---")


# ----------------------------------------------------------------------------
# Tableau de l'univers noté
# ----------------------------------------------------------------------------

st.markdown("### Univers des émetteurs notés")

tableau = resultats[resultats["Note finale"] != "N/A"][
    [
        "Emetteur",
        "Type",
        "Secteur",
        "Note 1",
        "Note 2",
        "Note 3",
        "Score",
        "Note finale",
    ]
].rename(
    columns={
        "Emetteur": "Émetteur",
        "Note finale": "Note",
    }
)


def couleur_note(v):
    metadata = RATING_META.get(v)
    if metadata:
        return f"color:{metadata['color']}; font-weight:700;"
    return ""


st.dataframe(
    tableau.style.map(
        couleur_note,
        subset=["Note 1", "Note 2", "Note 3", "Note"],
    ),
    use_container_width=True,
    hide_index=True,
)


# ----------------------------------------------------------------------------
# Export Excel
# ----------------------------------------------------------------------------

buffer = io.BytesIO()
export = resultats[resultats["Note finale"] != "N/A"].copy()

with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
    export.to_excel(writer, sheet_name="Notation", index=False)

    workbook = writer.book
    worksheet = writer.sheets["Notation"]

    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": NAVY,
            "font_color": "white",
            "border": 1,
        }
    )

    for index_colonne, nom_colonne in enumerate(export.columns):
        worksheet.write(0, index_colonne, nom_colonne, header_fmt)
        worksheet.set_column(
            index_colonne,
            index_colonne,
            max(14, len(str(nom_colonne)) + 2),
        )

st.download_button(
    "📥 Exporter la notation (Excel)",
    data=buffer.getvalue(),
    file_name="Notation_AFRICAPITAL.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.caption(
    "Modèle d'aide à la décision — ne se substitue ni à l'analyse fondamentale "
    "ni aux éléments qualitatifs : gouvernance, actionnariat et position concurrentielle."
)
